import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# URL = 'https://www.prokerala.com/travel/indian-railway/trains/abohar-jodhpur-express-14628.html'

# box1_data = box1_results.find('tbody')

# url = "trainSchedule.xlsx"
trainurl = "trainlist.xlsx"
wb = load_workbook(trainurl,data_only=True)

ws1 = wb['train']
ws2 = wb['schedule']
ws3 = wb['train_details']
# ws = wb['Schedule']
maxRows = ws1.max_row
maxRows2 = ws2.max_row
maxRows3 = ws3.max_row

allLinks = []

for a in range(2,maxRows+1):
	# allLinks.append([ws.cell(1,6)])
	temp = ws1.cell(a,6).value
	allLinks.append(temp)
allLinks1 = allLinks

# print(allLinks1)

f = open("completedData.txt", "r")
delList = f.read()
f.close()
y = delList.split(",")

# print(y)
while("" in y):
    y.remove("")

for each in y:
	allLinks.remove(each)

allLink = allLinks
newone = len(y)

index = 1 + maxRows2

serial = 0 + newone

id = 0

for x in allLink[:10]:
	id = id + 1
	serial = serial + 1
	newone = newone + 1
	print(newone,'\t',id)
	page = requests.get('https://www.prokerala.com/travel/indian-railway/trains/'+x)
	soup = BeautifulSoup(page.content, "html.parser")
	train_details = soup.find('table', {'class':'train-details'})

	if train_details:
		thead = train_details.find('thead').text
		
		tbody = train_details.find('tbody')
		tr1 = tbody.findAll('tr')[0]
		td11 = tr1.findAll('td')[0].text
		td12 = tr1.findAll('td')[1].text

		tr2 = tbody.findAll('tr')[1]
		td21 = tr2.findAll('td')[0].text
		td22 = tr2.findAll('td')[1].text

		tr3 = tbody.findAll('tr')[2]
		td31 = tr3.findAll('td')[0].text
		td32 = tr3.findAll('td')[1].text

		tr4 = tbody.findAll('tr')[3]
		td41 = tr4.findAll('td')[0].text
		td42 = tr4.findAll('td')[1].text

		tr5 = tbody.findAll('tr')[4]
		td51 = tr5.findAll('td')[0].text
		td52 = tr5.findAll('td')[1].text

		tr6 = tbody.findAll('tr')[5]
		td61 = tr6.findAll('td')[0].text
		td62 = tr6.findAll('td')[1].text

		ws3.cell(row=serial,column=1).value = serial
		ws3.cell(row=serial,column=2).value = thead
		ws3.cell(row=serial,column=3).value = td11
		ws3.cell(row=serial,column=4).value = td12
		ws3.cell(row=serial,column=5).value = td21
		ws3.cell(row=serial,column=6).value = td22
		ws3.cell(row=serial,column=7).value = td31
		ws3.cell(row=serial,column=8).value = td32
		ws3.cell(row=serial,column=9).value = td41
		ws3.cell(row=serial,column=10).value = td42
		ws3.cell(row=serial,column=11).value = td51
		ws3.cell(row=serial,column=12).value = td52
		ws3.cell(row=serial,column=13).value = td61
		ws3.cell(row=serial,column=14).value = td62

		train_table = soup.find('table', {'id':'train-table'})
		if train_table:
			train_table_tbody = train_table.find('tbody')
			tr = train_table_tbody.findAll('tr')

			for each in tr:
				index = index + 1

				if thead:
					ws2.cell(row=index,column=2).value = thead

				td0 = each.findAll('td')[0].text
				ws2.cell(row=index,column=1).value = td0
				td1 = each.findAll('td')[1].text
				ws2.cell(row=index,column=3).value = td1
				td2 = each.findAll('td')[2].text
				ws2.cell(row=index,column=4).value = td2
				td3 = each.findAll('td')[3].text
				ws2.cell(row=index,column=5).value = td3
				td4 = each.findAll('td')[4].text
				ws2.cell(row=index,column=6).value = td4
				td5 = each.findAll('td')[5].text
				ws2.cell(row=index,column=7).value = td5
				td6 = each.findAll('td')[6].text
				ws2.cell(row=index,column=8).value = td6
				td7 = each.findAll('td')[7].text
				ws2.cell(row=index,column=9).value = td7

	else:
		f = open("unavailableLink.txt", "a")
		f.write(x+',')
		f.close()

wb.save("trainlist.xlsx")
f = open("completedData.txt", "a")
f.write(x+',')
f.close()
