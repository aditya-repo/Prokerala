import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

URL = 'https://www.prokerala.com/travel/indian-railway/trains/'
page = requests.get(URL)
soup = BeautifulSoup(page.content, "html.parser")
alphabetTrain = soup.findAll('div', {'class':'table-responsive'})
# print(alphabetTrain)

# box1_results = soup.find(id="alpha-A")

xurl = "trainlist.xlsx"
wb = load_workbook(xurl,data_only=True)
ws = wb['train']

index = 0
for box1_results in alphabetTrain:
	box1_data = box1_results.find('tbody')
	table_row = box1_data.findAll('tr')
	for everytd in table_row:
		index = index + 1
		sr = everytd.findAll('td')[0].text
		ws.cell(row=index+1,column=1).value = sr
		train_no = everytd.findAll('td')[1].text
		ws.cell(row=index+1,column=2).value = train_no
		train_name = everytd.findAll('td')[2].text
		ws.cell(row=index+1,column=3).value = train_name
		arrivefrom = everytd.findAll('td')[3].text
		ws.cell(row=index+1,column=4).value = arrivefrom
		to = everytd.findAll('td')[4].text
		ws.cell(row=index+1,column=5).value = to

		train_names = everytd.findAll('td')[2]
		train_names = train_names.find('a')
		train_url = train_names.get('href')
		ws.cell(row=index+1,column=6).value = train_url

		print(sr,'\t',train_no,'\t',train_name,'\t',arrivefrom,'\t',to,'\t',train_url)


	wb.save('trainlist.xlsx')
